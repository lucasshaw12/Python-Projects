#! python3
# Takes 2 int and a filename as command arguments = 'python blankRowInserter.py [row number] [blank rows to add]'
# Start at row N, add M black rows into the workbook

import openpyxl, sys

# Take a numberToMultiply from the command line
if len(sys.argv) > 1:
	rowToInsert = int(sys.argv[1])
	print(rowToInsert)
if len(sys.argv) > 2: 
	blankRows = int(sys.argv[2]) 
if len(sys.argv) > 3:
	newFilename = sys.argv[3]

# TO DO
# Create a xlsx file and populate with test text
wb = openpyxl.Workbook()
sheet = wb.active
for rowNum in range(15):
	sheet.cell(row=rowNum+1, column=1).value = 'test text'
wb.save('blankRow.xlsx')

# copy workbook into a new workbook
newBook = openpyxl.load_workbook('blankRow.xlsx')
sheet = newBook.active

# Add blank rows to the new workbook
for j in range(blankRows):
	sheet.cell(row=rowToInsert+j, column=1).value = ' '

# Save to a new workbook
newBook.save('blankRowDone.xlsx')
# newBook.save(newFilename) # Uses terminal value to name the save file