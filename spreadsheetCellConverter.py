#! python3
# Invert/transpose the row/column order of a spreadsheet
# i.e row 5, column 3 should = column 3, row 5

import openpyxl

# Read the workbook
wb = openpyxl.load_workbook('blankrow.xlsx')
sheet = wb.active

newWb = openpyxl.Workbook()
invertedSheet = newWb.active

# Read the contents of the workbook
for x in range(1, 100): # For a full sheet, use sheet.max_row & sheet.max_column
	for y in range(1, 100): 
		invertedSheet.cell(row=y, column=x).value = sheet.cell(row=x, column=y).value 

# Save as a different filename
newWb.save('convertedCells.xlsx')