#! python3
# Creates a times table using a number given within the terminal arg
# Example format = 'python MultiplicationTableMaker.py 31'

import sys, openpyxl
from openpyxl.styles import Font 

# Take a numberToMultiply from the command line
if len(sys.argv) > 1:
	numberToMultiply = int(sys.argv[1])

# Create new Excel Spreadsheet
wb = openpyxl.Workbook()
sheet = wb.active
boldFontObj = Font(bold=True)

# Create rows and headers
for i in range(1, numberToMultiply+ 1):
	sheet['A' + str(i+1)].value = i
	sheet.cell(row=1,column=i+1).value = i

	sheet['A' + str(i+1)].font = boldFontObj
	sheet.cell(row=1, column=i+1).font = boldFontObj

# Populate the multiplication table using a nested loop to iterate over the cells
for j in range(1, numberToMultiply + 1):
	for k in range(1, numberToMultiply + 1):
		sheet.cell(row=j+1, column=k+1).value = j*k

# Save the workbook
wb.save('MultiplicationTable.xlsx')